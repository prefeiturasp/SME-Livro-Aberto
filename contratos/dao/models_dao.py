from django.db import IntegrityError, transaction

from openpyxl import load_workbook

from contratos.models import (
    CategoriaContrato,
    CategoriaContratoFromTo,
    ContratoRaw,
    Fornecedor,
    EmpenhoSOFCache,
    EmpenhoSOFCacheTemp,
    EmpenhoSOFFailedAPIRequest,
    ExecucaoContrato,
    ModalidadeContrato,
    ObjetoContrato)


class EmpenhosSOFCacheDao:

    def __init__(self):
        self.model = EmpenhoSOFCache

    def get_all(self):
        return self.model.objects.all()

    def filter_by_ano_exercicio_and_categoria(self, year):
        """
        Este método é utilizado para filtrar e agrupar os dados de contratos 
        por ano para gerar o arquivo para download.
        """
        return self.model.objects.filter(
            anoExercicioContrato=year,
            execucaocontrato__categoria__isnull=False) \
            .order_by('anoEmpenho', 'codContrato')

    def create(self, data):
        return self.model.objects.create(**data)

    def count_all(self):
        return self.model.objects.count()

    def create_from_temp_table_obj(self, empenho_temp):
        empenho = self.model()
        for field in empenho_temp._meta.fields:
            if field.primary_key is True:
                continue
            setattr(empenho, field.name, getattr(empenho_temp, field.name))
        empenho.save()
        return empenho


class EmpenhosSOFCacheTempDao:

    def __init__(self):
        self.model = EmpenhoSOFCacheTemp

    def get_all(self):
        return self.model.objects.all()

    def create(self, data):
        return self.model.objects.create(**data)

    def count_all(self):
        return self.model.objects.count()

    def delete(self, obj):
        obj.delete()

    def erase_all(self):
        self.model.objects.all().delete()


class EmpenhosFailedRequestsDao:

    def __init__(self):
        self.model = EmpenhoSOFFailedAPIRequest

    def create(self, cod_contrato, ano_exercicio, ano_empenho, error_code):
        return self.model.objects.create(
            cod_contrato=cod_contrato, ano_exercicio=ano_exercicio,
            ano_empenho=ano_empenho, error_code=error_code)

    def get_all(self):
        return self.model.objects.all()

    def count_all(self):
        return self.model.objects.count()

    def delete(self, obj):
        obj.delete()

    def erase_all(self):
        self.model.objects.all().delete()


class ContratosRawDao:

    def __init__(self):
        self.model = ContratoRaw

    def get_all(self):
        return self.model.objects.all().order_by("anoExercicioContrato",
                                                 "codContrato")

    def get(self, **data):
        return self.model.objects.get(**data)


class ExecucoesContratosDao:

    def __init__(self):
        self.model = ExecucaoContrato

    def create(self, **data):
        return self.model.objects.create(**data)

    def filter_by_indexer(self, indexer):
        return self.model.objects.filter(empenho_indexer=indexer)

    def update_with(self, execucao, **data):
        for field, value in data.items():
            setattr(execucao, field, value)
        return execucao.save()

    def erase_all(self):
        self.model.objects.all().delete()


class ModalidadesContratosDao:

    def __init__(self):
        self.model = ModalidadeContrato

    def get_or_create(self, **data):
        return self.model.objects.get_or_create(**data)


class ObjetosContratosDao:

    def __init__(self):
        self.model = ObjetoContrato

    def get_or_create(self, **data):
        return self.model.objects.get_or_create(**data)


class FornecedoresDao:

    def __init__(self):
        self.model = Fornecedor

    def get_or_create(self, **data):
        return self.model.objects.get_or_create(**data)


class CategoriasContratosFromToDao:

    def __init__(self):
        self.model = CategoriaContratoFromTo

    def get_all(self):
        return self.model.objects.all()

    def extract_spreadsheet(self, ssheet_obj):
        if ssheet_obj.extracted:
            return

        filepath = ssheet_obj.spreadsheet.path
        wb = load_workbook(filepath)
        ws = wb.worksheets[0]

        row = 2
        row_is_valid = True
        added = []
        not_added = []
        while row_is_valid:
            indexer = ws['a' + str(row)].value
            if not indexer:
                row_is_valid = False
                continue
            categoria_name = ws['b' + str(row)].value.strip()
            categoria_desc = ws['c' + str(row)].value.strip()

            cat_ft = CategoriaContratoFromTo()
            cat_ft.indexer = indexer
            cat_ft.categoria_name = categoria_name
            cat_ft.categoria_desc = categoria_desc
            try:
                with transaction.atomic():
                    cat_ft.save()
                added.append(cat_ft.indexer)
            except IntegrityError:
                not_added.append(cat_ft.indexer)

            row += 1

        ssheet_obj.added_fromtos = added
        ssheet_obj.not_added_fromtos = not_added
        ssheet_obj.extracted = True
        ssheet_obj.save()
        return added, not_added


class CategoriasContratosDao:

    def __init__(self):
        self.model = CategoriaContrato

    def get_or_create(self, **data):
        return self.model.objects.get_or_create(**data)

    def update_with(self, categoria, **data):
        for field, value in data.items():
            setattr(categoria, field, value)
        return categoria.save()
