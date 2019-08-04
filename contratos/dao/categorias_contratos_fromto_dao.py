from django.db import IntegrityError, transaction
from openpyxl import load_workbook

from contratos.models import CategoriaContratoFromTo


def extract_spreadsheet(ssheet_obj):
    if ssheet_obj.extracted:
        return

    filepath = ssheet_obj.spreadsheet.path
    wb = load_workbook(filepath)
    ws = wb['Sheet1']

    row = 2
    row_is_valid = True
    added = []
    not_added = []
    while row_is_valid:
        indexer = ws['a' + str(row)].value
        if not indexer:
            row_is_valid = False
            continue
        categoria_name = ws['b' + str(row)].value
        categoria_desc = ws['c' + str(row)].value

        dot = CategoriaContratoFromTo()
        dot.indexer = indexer
        dot.categoria_name = categoria_name
        dot.categoria_desc = categoria_desc
        try:
            with transaction.atomic():
                dot.save()
            added.append(dot.indexer)
        except IntegrityError:
            # add to dot.not_added
            not_added.append(dot.indexer)

        row += 1

    ssheet_obj.added_fromtos = added
    ssheet_obj.not_added_fromtos = not_added
    ssheet_obj.extracted = True
    ssheet_obj.save()
    return added, not_added
