from collections import namedtuple
from datetime import date
from itertools import groupby

from django.db import IntegrityError, transaction
from openpyxl import load_workbook

from regionalizacao.models import (
    DistritoZonaFromTo,
    EtapaTipoEscolaFromTo,
    PtrfFromTo,
    UnidadeRecursosFromTo,
    PtrfFromToSpreadsheet,
    UnidadeRecursosFromToSpreadsheet,
    Dre,
    TipoEscola,
    Distrito,
    Escola,
    EscolaInfo,
    Recurso,
    Subgrupo,
    Budget,
    Grupo,
)


SheetColumn = namedtuple('SheetColumn', ['name', 'letter'])


class FromToDao:

    def extract_spreadsheet(self, sheet):
        filepath = sheet.spreadsheet.path
        wb = load_workbook(filepath)
        ws = wb.worksheets[0]

        row = 2
        row_is_valid = True
        added = []
        updated = []
        while row_is_valid:
            ft_key_name = self.sheet_columns[0].name
            ft_key = ws[self.sheet_columns[0].letter + str(row)].value
            if not ft_key:
                row_is_valid = False
                continue

            ft = self.model()
            for column in self.sheet_columns:
                value = ws[column.letter + str(row)].value
                setattr(ft, column.name, value)

            # setting year when applies
            year = getattr(sheet, 'year', None)
            if year:
                ft.year = year

            # saving fromto object
            try:
                with transaction.atomic():
                    ft.save()
                added.append(ft_key)
            except IntegrityError:
                old_ft = self.model.objects.get(**{ft_key_name: ft_key})
                old_ft.delete()
                ft.save()
                updated.append(ft_key)

            row += 1

        sheet.added_fromtos = added
        sheet.updated_fromtos = updated
        sheet.extracted = True
        sheet.save()
        return added, updated


class PtrfFromToDao(FromToDao):

    def __init__(self):
        self.model = PtrfFromTo
        self.sheet_columns = [
            SheetColumn('codesc', 'a'),
            SheetColumn('vlrepasse', 'b'),
        ]

    def extract_spreadsheet(self, sheet):
        year_fromtos = self.model.objects.filter(year=sheet.year)
        year_fromtos.delete()
        return super().extract_spreadsheet(sheet)

    def get_all(self):
        return self.model.objects.all().order_by('codesc')


class DistritoZonaFromToDao(FromToDao):

    def __init__(self):
        self.model = DistritoZonaFromTo
        self.sheet_columns = [
            SheetColumn('coddist', 'a'),
            SheetColumn('zona', 'b'),
        ]

    def get_all(self):
        return self.model.objects.all().order_by('coddist')


class EtapaTipoEscolaFromToDao(FromToDao):

    def __init__(self):
        self.model = EtapaTipoEscolaFromTo
        self.sheet_columns = [
            SheetColumn('tipoesc', 'a'),
            SheetColumn('desctipoesc', 'b'),
            SheetColumn('etapa', 'c'),
        ]

    def get_all(self):
        return self.model.objects.all().order_by('tipoesc')


class UnidadeRecursosFromToDao(FromToDao):

    def __init__(self):
        self.model = UnidadeRecursosFromTo
        self.sheet_columns = [
            SheetColumn('codesc', 'a'),
            SheetColumn('grupo', 'b'),
            SheetColumn('subgrupo', 'c'),
            SheetColumn('valor', 'd'),
            SheetColumn('label', 'e'),
        ]

    def extract_spreadsheet(self, sheet):
        year_fromtos = self.model.objects.filter(year=sheet.year)
        year_fromtos.delete()
        return super().extract_spreadsheet(sheet)

    def get_all(self):
        return self.model.objects.all().order_by('year', 'codesc')


class FromToSpreadsheetDao:

    def get_sheets_to_be_extracted(self):
        return self.model.objects.filter(extracted=False) \
            .order_by('created_at')

    def get_years_to_be_updated(self):
        sheets = self.get_sheets_to_be_extracted()
        years = sheets.order_by('year').values_list('year', flat=True) \
            .distinct()
        return list(years)

    def extract_new_spreadsheets(self):
        sheets = self.get_sheets_to_be_extracted()
        for sheet in sheets:
            sheet.extract_data()

    def get_last_created_at(self):
        sheet = self.model.objects.filter(extracted=True) \
            .order_by('-created_at').first()
        if sheet:
            return sheet.created_at.date()


class PtrfFromToSpreadsheetDao(FromToSpreadsheetDao):

    def __init__(self):
        self.model = PtrfFromToSpreadsheet


class UnidadeRecursosFromToSpreadsheetDao(FromToSpreadsheetDao):

    def __init__(self):
        self.model = UnidadeRecursosFromToSpreadsheet


class DreDao:

    def __init__(self):
        self.model = Dre

    def update_or_create(self, code, name):
        try:
            with transaction.atomic():
                dre = self.create(code=code, name=name)
            created = True
        except IntegrityError:
            dre = self.update(code=code, name=name)
            created = False

        return dre, created

    def create(self, code, name):
        return self.model.objects.create(code=code, name=name)

    def update(self, code, name):
        dre = self.model.objects.get(code=code)
        dre.name = name
        dre.save()
        return dre


class TipoEscolaDao:

    def __init__(self):
        self.model = TipoEscola

    def get(self, code):
        try:
            return self.model.objects.get(code=code)
        except self.model.DoesNotExist:
            return None

    def get_or_create(self, code):
        return self.model.objects.get_or_create(code=code)


class DistritoDao:

    def __init__(self):
        self.model = Distrito

    def get(self, coddist):
        try:
            return self.model.objects.get(coddist=coddist)
        except self.model.DoesNotExist:
            return None

    def get_or_create(self, coddist, name):
        return self.model.objects.get_or_create(
            coddist=coddist, defaults={'name': name})


class EscolaDao:

    def __init__(self):
        self.model = Escola
        self.info_dao = EscolaInfoDao()
        self.dre_dao = DreDao()
        self.tipo_dao = TipoEscolaDao()
        self.distrito_dao = DistritoDao()

    def get(self, codesc, year):
        try:
            return self.model.objects.get(codesc=codesc, year=year)
        except self.model.DoesNotExist:
            return None

    def get_or_create(self, codesc):
        return self.model.objects.get_or_create(codesc=codesc)

    def update_or_create(self, **kwargs):
        escola, created = self.get_or_create(codesc=kwargs['codesc'])

        dre, _ = self.dre_dao.update_or_create(
            code=kwargs['dre'], name=kwargs['diretoria'].strip())
        tipo, _ = self.tipo_dao.get_or_create(code=kwargs['tipoesc'])
        distrito, _ = self.distrito_dao.get_or_create(
            coddist=kwargs['coddist'], name=kwargs['distrito'].strip())

        escola_info = dict(
            escola_id=escola.id,
            year=date.today().year,
            dre=dre,
            tipoesc=tipo,
            distrito=distrito,
            nomesc=kwargs['nomesc'].strip(),
            endereco=kwargs['endereco'].strip(),
            numero=kwargs['numero'].strip(),
            bairro=kwargs['bairro'].strip(),
            cep=kwargs['cep'],
            rede=kwargs['rede'],
            latitude=kwargs['latitude'],
            longitude=kwargs['longitude'],
            total_vagas=kwargs['total_vagas'],
        )

        self.info_dao.update_or_create(**escola_info)

        return escola, created

    def create_for_previous_year(self, **kwargs):
        escola, _ = self.get_or_create(codesc=kwargs['codesc'])

        dre, _ = self.dre_dao.update_or_create(
            code=kwargs['dre'], name=kwargs['diretoria'].strip())
        tipo, _ = self.tipo_dao.get_or_create(code=kwargs['tipoesc'])
        distrito, _ = self.distrito_dao.get_or_create(
            coddist=kwargs['coddist'], name=kwargs['distrito'].strip())

        escola_info = dict(
            escola_id=escola.id,
            year=kwargs['year'],
            dre=dre,
            tipoesc=tipo,
            distrito=distrito,
            nomesc=kwargs['nomesc'].strip(),
            endereco=kwargs['endereco'].strip(),
            numero=kwargs['numero'].strip(),
            bairro=kwargs['bairro'].strip(),
            cep=kwargs['cep'],
            rede=kwargs['rede'],
            latitude=kwargs['latitude'],
            longitude=kwargs['longitude'],
            total_vagas=kwargs['total_vagas'],
        )

        _, created = self.info_dao.get_or_create(**escola_info)

        return escola, created

    def create(self, **data):
        return self.model.objects.create(**data)


class EscolaInfoDao:

    def __init__(self):
        self.model = EscolaInfo

    def get(self, escola_id, year):
        try:
            return self.model.objects.get(escola__id=escola_id, year=year)
        except self.model.DoesNotExist:
            return None

    def update_or_create(self, **data):
        try:
            with transaction.atomic():
                info = self.create(**data)
            created = True
        except IntegrityError:
            info = self.update(**data)
            created = False

        return info, created

    def get_or_create(self, **data):
        escola_id = data.pop('escola_id')
        year = data.pop('year')
        return self.model.objects.get_or_create(
            escola_id=escola_id, year=year, defaults=data)

    def create(self, **data):
        return self.model.objects.create(**data)

    def update(self, **data):
        escola_id = data.pop('escola_id')
        year = data.pop('year')
        info = self.get(escola_id=escola_id, year=year)
        if not info:
            return None

        for field_name, value in data.items():
            setattr(info, field_name, value)
        info.save()
        return info


class BudgetDao:

    def __init__(self):
        self.model = Budget
        self.escola_dao = EscolaDao()

    def get_all(self):
        return self.model.objects.all()

    def get(self, escola_id, year):
        try:
            return self.model.objects.get(escola_id=escola_id, year=year)
        except self.model.DoesNotExist:
            return None

    def get_or_create(self, codesc, year):
        escola, _ = self.escola_dao.get_or_create(codesc=codesc)
        return self.model.objects.get_or_create(escola=escola, year=year)

    def update_or_create(self, **data):
        codesc = data.pop('codesc')
        escola, _ = self.escola_dao.get_or_create(codesc=codesc)
        data['escola_id'] = escola.id
        try:
            with transaction.atomic():
                info = self.create(**data)
            created = True
        except IntegrityError:
            info = self.update(**data)
            created = False

        return info, created

    def create(self, **data):
        return self.model.objects.create(**data)

    def update(self, **data):
        escola_id = data.pop('escola_id')
        year = data.pop('year')
        budget = self.get(escola_id=escola_id, year=year)

        for field_name, value in data.items():
            setattr(budget, field_name, value)
        budget.save()
        return budget

    def build_recursos_data(self, budget):
        qs = budget.recursos.all().order_by('subgrupo__grupo__name')
        total = 0

        grupos, total_g = self._build_grupos_data(qs)
        ptrf = budget.ptrf if budget.ptrf else 0
        total += total_g + ptrf
        recursos_dict = {
            'ptrf': ptrf,
            'grupos': grupos,
        }

        return recursos_dict, total

    def _build_grupos_data(self, recursos):
        grupos_list = []
        total = 0
        for grupo_name, recursos_g \
                in groupby(recursos, lambda r: r.subgrupo.grupo.name):
            recursos_g = list(recursos_g)

            subgrupos, total_s = self._build_subgrupos_data(recursos_g)
            grupo_dict = {
                "name": grupo_name,
                "total": total_s,
                "subgrupos": subgrupos,
            }
            grupos_list.append(grupo_dict)
            total += total_s

        grupos_list.sort(key=lambda g: g['total'], reverse=True)
        return grupos_list, total

    def _build_subgrupos_data(self, recursos):
        subgrupos_list = []
        total = 0
        if len(recursos) == 1:
            recurso = recursos[0]
            total = recurso.cost if recurso.cost else 0
            return subgrupos_list, total

        for recurso in recursos:
            cost = recurso.cost if recurso.cost else 0
            subgrupo_dict = {
                'name': recurso.subgrupo.name,
                'cost': cost,
                'amount': recurso.amount,
                'label': recurso.label,
            }
            subgrupos_list.append(subgrupo_dict)
            total += cost

        subgrupos_list.sort(key=lambda s: s['cost'], reverse=True)
        return subgrupos_list, total


class RecursoDao:

    def __init__(self):
        self.model = Recurso
        self.budget_dao = BudgetDao()
        self.subgrupo_dao = SubgrupoDao()

    def get(self, budget_id, subgrupo_id):
        try:
            return self.model.objects.get(budget_id=budget_id,
                                          subgrupo_id=subgrupo_id)
        except self.model.DoesNotExist:
            return None

    def update_or_create(self, codesc, year, grupo_name, subgrupo_name, valor,
                         label):
        budget, _ = self.budget_dao.get_or_create(codesc=codesc, year=year)
        subgrupo, _ = self.subgrupo_dao.get_or_create(name=subgrupo_name,
                                                      grupo_name=grupo_name)
        data = dict(budget_id=budget.id, subgrupo_id=subgrupo.id)
        if label == 'R$':
            data['cost'] = valor
        else:
            data['amount'] = valor
            data['label'] = label

        try:
            with transaction.atomic():
                recurso = self.create(**data)
            created = True
        except IntegrityError:
            recurso = self.update(**data)
            created = False

        return recurso, created

    def create(self, **data):
        return self.model.objects.create(**data)

    def update(self, **data):
        budget_id = data.pop('budget_id')
        subgrupo_id = data.pop('subgrupo_id')
        recurso = self.get(budget_id=budget_id, subgrupo_id=subgrupo_id)

        for field_name, value in data.items():
            setattr(recurso, field_name, value)
        recurso.save()
        return recurso


class SubgrupoDao:

    def __init__(self):
        self.model = Subgrupo
        self.grupo_dao = GrupoDao()

    def get_or_create(self, name, grupo_name):
        grupo, _ = self.grupo_dao.get_or_create(name=grupo_name)
        return self.model.objects.get_or_create(name=name, grupo=grupo)


class GrupoDao:

    def __init__(self):
        self.model = Grupo

    def get_or_create(self, name):
        return self.model.objects.get_or_create(name=name)
