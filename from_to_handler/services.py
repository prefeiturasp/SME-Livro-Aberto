from django.db import IntegrityError, transaction
from openpyxl import load_workbook

from from_to_handler.models import DotacaoFromTo


def extract_dotacao_fromto_spreadsheet(ssheet_obj):
    filepath = ssheet_obj.spreadsheet.path
    wb = load_workbook(filepath)
    ws = wb['Planilha1']

    row = 2
    row_is_valid = True
    added = []
    not_added = []
    while row_is_valid:
        indexer = ws['a' + str(row)].value
        if not indexer:
            row_is_valid = False
            continue
        grupo_code = int(ws['d' + str(row)].value)
        grupo_desc = ws['e' + str(row)].value
        subgrupo_code = int(ws['b' + str(row)].value.split('.')[1])
        subgrupo_desc = ws['c' + str(row)].value

        dot = DotacaoFromTo()
        dot.indexer = indexer
        dot.grupo_code = grupo_code
        dot.grupo_desc = grupo_desc
        dot.subgrupo_code = subgrupo_code
        dot.subgrupo_desc = subgrupo_desc
        try:
            with transaction.atomic():
                dot.save()
            added.append(dot.indexer)
        except IntegrityError:
            # add to dot.not_added
            not_added.append(dot.indexer)

        row += 1

    return added, not_added
