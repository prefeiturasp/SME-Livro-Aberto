# # This script was meant to be runned only once during the development.
# # It shouldn't be used anymore. The data gathered by this script is in
# # a json file (from_to_handler/migrations/data/from_to.json) that is loaded
# # by a datamigration (migration 0009 in from_to_handler).


import datetime

from django.db import IntegrityError
from openpyxl import load_workbook

from from_to_handler.models import (
    Deflator,
    DotacaoFromTo,
    FonteDeRecursoFromTo,
    GNDFromTo,
    SubelementoFromTo,
)


def import_fontes_de_recurso():
    filepath = './from_to_handler/de-paras/DE-PARA Fontes de Recurso.xlsx'
    wb = load_workbook(filepath)
    ws = wb['Plan1']

    row = 2
    row_is_valid = True
    while row_is_valid:
        try:
            code = int(ws['a' + str(row)].value)
        except TypeError:
            row_is_valid = False
            continue
        name = ws['b' + str(row)].value
        grupo_code = int(ws['c' + str(row)].value)
        grupo_name = ws['d' + str(row)].value

        fdr = FonteDeRecursoFromTo()
        fdr.code = code
        fdr.name = name
        fdr.grupo_code = grupo_code
        fdr.grupo_name = grupo_name
        fdr.save()

        row += 1


def import_subelementos():
    filepath = './from_to_handler/de-paras/DE-PARA Sub-elementos.xlsx'
    wb = load_workbook(filepath)
    ws = wb['Plan1']

    row = 2
    row_is_valid = True
    while row_is_valid:
        code = ws['a' + str(row)].value
        if not code:
            row_is_valid = False
            continue
        desc = ws['b' + str(row)].value
        new_code = int(ws['c' + str(row)].value)
        new_name = ws['d' + str(row)].value

        se = SubelementoFromTo()
        se.code = code
        se.desc = desc
        se.new_code = new_code
        se.new_name = new_name
        se.save()

        row += 1


def import_dotacoes():
    filepath = ('./from_to_handler/de-paras/'
                'DE-PARA Dotações Subgrupos Grupos.xlsx')
    wb = load_workbook(filepath)
    ws = wb['De Para Final']

    row = 2
    row_is_valid = True
    while row_is_valid:
        indexer = ws['b' + str(row)].value
        if not indexer:
            row_is_valid = False
            continue
        grupo_code = int(ws['f' + str(row)].value)
        grupo_desc = ws['g' + str(row)].value
        subgrupo_code = int(ws['d' + str(row)].value.split('.')[1])
        subgrupo_desc = ws['e' + str(row)].value

        dot = DotacaoFromTo()
        dot.indexer = indexer
        dot.grupo_code = grupo_code
        dot.grupo_desc = grupo_desc
        dot.subgrupo_code = subgrupo_code
        dot.subgrupo_desc = subgrupo_desc
        try:
            dot.save()
        except IntegrityError:
            pass

        row += 1


def import_gnd():
    filepath = ('./from_to_handler/de-paras/'
                'DE-PARA Grupos de Despesa e Elementos.xlsx')
    wb = load_workbook(filepath)
    ws = wb['Plan1']

    row = 2
    row_is_valid = True
    while row_is_valid:
        try:
            gnd_code = int(ws['e' + str(row)].value)
        except TypeError:
            row_is_valid = False
            continue
        gnd_desc = ws['a' + str(row)].value
        elemento_code = int(ws['d' + str(row)].value)
        elemento_desc = ws['c' + str(row)].value
        new_gnd_code = int(ws['f' + str(row)].value)
        new_gnd_desc = ws['b' + str(row)].value

        gnd = GNDFromTo()
        gnd.gnd_code = gnd_code
        gnd.gnd_desc = gnd_desc
        gnd.elemento_code = elemento_code
        gnd.elemento_desc = elemento_desc
        gnd.new_gnd_code = new_gnd_code
        gnd.new_gnd_desc = new_gnd_desc
        gnd.save()

        row += 1


def import_deflator():
    filepath = ('./from_to_handler/de-paras/'
                'Deflator Setembro 2018.xlsx')
    wb = load_workbook(filepath, data_only=True)
    ws = wb['Anual']

    row = 2
    row_is_valid = True
    while row_is_valid:
        year = ws['a' + str(row)].value
        if not year:
            row_is_valid = False
            continue
        index_number = ws['b' + str(row)].value
        variation_percent = ws['c' + str(row)].value

        defl = Deflator()
        defl.year = datetime.date(int(year), 1, 1)
        defl.index_number = index_number
        defl.variation_percent = variation_percent * 100
        defl.save()

        row += 1


def run():
    # import_fontes_de_recurso()
    # import_subelementos()
    # import_dotacoes()
    # import_gnd()
    # import_deflator()
    pass
