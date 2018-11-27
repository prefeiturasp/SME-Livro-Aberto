# # This script was meant to be runned only once during the development.
# # It shouldn't be used anymore. The data gathered by this script is in
# # a json file (budget_execution/data/budget_execution_models_data.json) that
# # needs to be loaded manually with `manage.py loaddata`

import datetime

from decimal import Decimal
from typing import NamedTuple

from django.apps import apps
from openpyxl import load_workbook

from budget_execution.models import Execucao


class OtherColumn(NamedTuple):
    name: str
    column: str


class ColumnsLetters(NamedTuple):
    id: str
    description: str = None
    other_col: OtherColumn = None


FK_OBJECTS_COLS = {
    # orcamento spreadsheet
    "Orgao": ColumnsLetters('H', 'J', OtherColumn('initials', 'I')),
    "ProjetoAtividade": ColumnsLetters('U', 'V', OtherColumn('type', 'S')),
    "Categoria": ColumnsLetters('Y', 'Z'),
    "Gnd": ColumnsLetters('AA', 'AB'),
    "Modalidade": ColumnsLetters('AC', 'AD'),
    "Elemento": ColumnsLetters('AE'),
    "FonteDeRecurso": ColumnsLetters('AF', 'AG'),
    "Subfuncao": ColumnsLetters('O', 'P'),
    "Programa": ColumnsLetters('Q', 'R'),

    # empenhos spreadsheet
    "Subelemento": ColumnsLetters('O', 'AE'),
}

EMPENHOS_COLS = {
    "Orgao": 'L',
    "ProjetoAtividade": 'N',
    "Categoria": 'C',
    "Gnd": 'I',
    "Modalidade": 'K',
    "Elemento": 'D',
    "FonteDeRecurso": 'G',
    # "Subfuncao": 'P',
    # "Programa": 'M',
}


def import_fk_object(ws, row, name):
    id_col = FK_OBJECTS_COLS[name].id
    desc_col = FK_OBJECTS_COLS[name].description
    other_col_tuple = FK_OBJECTS_COLS[name].other_col

    row = str(row)
    id = int(ws[id_col + row].value)
    if name == "Orgao" and id != 16:
        return None

    Model = apps.get_model('budget_execution', name)
    if Model.objects.filter(id=id).exists():
        return id

    obj = Model()
    obj.id = id

    if desc_col:
        description = ws[desc_col + row].value
        obj.description = description.strip()

    if other_col_tuple:
        other_col = other_col_tuple.column
        other_col_name = other_col_tuple.name
        other_col_value = ws[other_col + row].value
        setattr(obj, other_col_name, other_col_value.strip())

    obj.save()

    return id


def get_empenho_fk_object_id(ws, row, name):
    return int(ws[EMPENHOS_COLS[name] + str(row)].value)


def import_empenhos_csv():
    filepath = ('./budget_execution/data/'
                'empenhos-1541109573297.xlsx')
    wb = load_workbook(filepath)
    ws = wb['data']

    row = 2
    row_is_valid = True
    dotacoes_not_found = 0
    subelemento_added_count = 0
    empenho_updated_count = 0
    new_dotacoes = 0

    while row_is_valid:
        print(row)
        try:
            year = int(ws['b' + str(row)].value)
        except TypeError:
            row_is_valid = False
            continue

        orgao_id = get_empenho_fk_object_id(ws, row, 'Orgao')
        projeto_id = get_empenho_fk_object_id(ws, row, 'ProjetoAtividade')
        categoria_id = get_empenho_fk_object_id(ws, row, 'Categoria')
        gnd_id = get_empenho_fk_object_id(ws, row, 'Gnd')
        modalidade_id = get_empenho_fk_object_id(ws, row, 'Modalidade')
        elemento_id = get_empenho_fk_object_id(ws, row, 'Elemento')
        fonte_id = get_empenho_fk_object_id(ws, row, 'FonteDeRecurso')

        subelemento_id = import_fk_object(ws, row, 'Subelemento')
        empenhado = ws['al' + str(row)].value
        empenhado = Decimal(empenhado)

        execs = Execucao.objects.filter_by_indexer((
            f'{year}.{orgao_id}.{projeto_id}.{categoria_id}.{gnd_id}.'
            f'{modalidade_id}.{elemento_id}.{fonte_id}'))

        if len(execs) == 0:
            print("dotação not found")
            dotacoes_not_found += 1
            row += 1
            continue

        try:
            execucao = execs.get(subelemento_id=subelemento_id)
            execucao.empenhado_liquido += empenhado
            empenho_updated_count += 1
            print("empenhado updated")
            execucao.save()
            row += 1
            continue
        except Execucao.DoesNotExist:
            pass

        execs_without_subelemento = execs.filter(subelemento_id=None)

        if execs_without_subelemento:
            execucao = execs_without_subelemento[0]
            execucao.subelemento_id = subelemento_id
            subelemento_added_count += 1
            print("subelemento added")
        else:
            execucao = execs[0]
            execucao.pk = None
            execucao.subelemento_id = subelemento_id
            new_dotacoes += 1
            print("new execution")

        execucao.empenhado_liquido = empenhado

        execucao.save()
        row += 1

    print(f'new dotacoes: {new_dotacoes}')
    print(f'subelemento_added_count: {subelemento_added_count}')
    print(f'empenho_updated_count: {empenho_updated_count}')


def import_orcamento_csv():
    filepath = ('./budget_execution/data/'
                'orcamento-1541109528703.xlsx')
    wb = load_workbook(filepath)
    ws = wb['data']

    row = 2
    row_is_valid = True
    while row_is_valid:
        try:
            year = int(ws['d' + str(row)].value)
        except TypeError:
            row_is_valid = False
            continue

        orgao_id = import_fk_object(ws, row, 'Orgao')
        if not orgao_id:
            row += 1
            continue

        projeto_id = import_fk_object(ws, row, 'ProjetoAtividade')
        categoria_id = import_fk_object(ws, row, 'Categoria')
        gnd_id = import_fk_object(ws, row, 'Gnd')
        modalidade_id = import_fk_object(ws, row, 'Modalidade')
        elemento_id = import_fk_object(ws, row, 'Elemento')
        fonte_id = import_fk_object(ws, row, 'FonteDeRecurso')
        subfuncao_id = import_fk_object(ws, row, 'Subfuncao')
        programa_id = import_fk_object(ws, row, 'Programa')

        orcado = ws['ai' + str(row)].value

        try:
            execucao = Execucao.objects.get(
                year=datetime.date(year, 1, 1),
                orgao_id=orgao_id,
                projeto_id=projeto_id,
                categoria_id=categoria_id,
                gnd_id=gnd_id,
                modalidade_id=modalidade_id,
                elemento_id=elemento_id,
                fonte_id=fonte_id)
        except Execucao.DoesNotExist:
            execucao = Execucao()
            execucao.year = datetime.date(year, 1, 1)
            execucao.orgao_id = orgao_id
            execucao.projeto_id = projeto_id
            execucao.categoria_id = categoria_id
            execucao.gnd_id = gnd_id
            execucao.modalidade_id = modalidade_id
            execucao.elemento_id = elemento_id
            execucao.fonte_id = fonte_id

        execucao.subfuncao_id = subfuncao_id
        execucao.programa_id = programa_id

        if execucao.orcado_atualizado:
            execucao.orcado_atualizado += Decimal(orcado)
        else:
            execucao.orcado_atualizado = orcado

        execucao.save()
        print("Saved execução " + str(execucao.id))

        row += 1


def run():
    # import_orcamento_csv()
    # import_empenhos_csv()
    pass
